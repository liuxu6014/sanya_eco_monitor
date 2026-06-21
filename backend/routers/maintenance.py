"""设备运维 / 掉线统计接口。

供"专项分析 → 设备运维"板块使用：掉线明细清单、次数与时长统计、Excel 导出。
"""

from __future__ import annotations

from datetime import datetime, timedelta
from io import BytesIO
from urllib.parse import quote

from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import Response
from sqlalchemy.ext.asyncio import AsyncSession

from database import get_db
from services.device_uptime import (
    build_device_registry,
    compute_outage_report,
    get_device_by_key,
)
from time_utils import cn_now_naive

router = APIRouter(prefix="/api/maintenance", tags=["设备运维"])

_DATE_FMT = "%Y-%m-%d"


def _parse_range(start: str | None, end: str | None) -> tuple[datetime, datetime]:
    now = cn_now_naive()
    if end:
        try:
            end_dt = datetime.strptime(end, _DATE_FMT).replace(hour=23, minute=59, second=59)
        except ValueError as exc:
            raise HTTPException(status_code=400, detail="end 日期格式应为 YYYY-MM-DD") from exc
    else:
        end_dt = now
    if start:
        try:
            start_dt = datetime.strptime(start, _DATE_FMT).replace(hour=0, minute=0, second=0)
        except ValueError as exc:
            raise HTTPException(status_code=400, detail="start 日期格式应为 YYYY-MM-DD") from exc
    else:
        start_dt = (end_dt - timedelta(days=30)).replace(hour=0, minute=0, second=0)
    if end_dt <= start_dt:
        raise HTTPException(status_code=400, detail="end 必须晚于 start")
    return start_dt, end_dt


def _format_duration(seconds: int) -> str:
    seconds = max(0, int(seconds))
    days, rem = divmod(seconds, 86400)
    hours, rem = divmod(rem, 3600)
    minutes, _ = divmod(rem, 60)
    parts: list[str] = []
    if days:
        parts.append(f"{days}天")
    if hours:
        parts.append(f"{hours}小时")
    if minutes or not parts:
        parts.append(f"{minutes}分钟")
    return "".join(parts)


# 虫情、孢子约一天一报，时长以"天"显示更直观。
_LOW_FREQ_TYPES = {"insect", "spore"}

# 同一台径流站内置雨量计、用同一编码上报两路数据，导出时也需用类型区分。
_TYPE_LABELS = {"insect": "虫情", "spore": "孢子", "water": "水质", "runoff": "径流", "rain": "雨量"}


def _device_label(name: str, device_type: str) -> str:
    tag = _TYPE_LABELS.get(device_type)
    return f"{name}（{tag}）" if tag else name


def _format_duration_by_type(seconds: int, device_type: str) -> str:
    if device_type in _LOW_FREQ_TYPES:
        # 低频设备精确到小时即可：天+小时，不足 1 小时退回分钟。
        total = max(0, int(seconds))
        days, rem = divmod(total, 86400)
        hours = rem // 3600
        if days or hours:
            return (f"{days}天" if days else "") + f"{hours}小时"
        return f"{total // 60}分钟"
    return _format_duration(seconds)


@router.get("/devices")
async def list_devices():
    """设备下拉列表（筛选用）。"""
    return {
        "data": [
            {"key": d.key, "name": d.name, "type": d.type}
            for d in build_device_registry()
        ]
    }


@router.get("/outages")
async def list_outages(
    device: str | None = Query(default=None, description="设备 key，留空表示全部设备"),
    start: str | None = Query(default=None, description="开始日期 YYYY-MM-DD"),
    end: str | None = Query(default=None, description="结束日期 YYYY-MM-DD"),
    db: AsyncSession = Depends(get_db),
):
    if device and get_device_by_key(device) is None:
        raise HTTPException(status_code=404, detail="未知设备")
    start_dt, end_dt = _parse_range(start, end)
    report = await compute_outage_report(db, start=start_dt, end=end_dt, device_key=device)
    return {"data": report}


@router.get("/outages/export")
async def export_outages(
    device: str | None = Query(default=None),
    start: str | None = Query(default=None),
    end: str | None = Query(default=None),
    db: AsyncSession = Depends(get_db),
):
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    if device and get_device_by_key(device) is None:
        raise HTTPException(status_code=404, detail="未知设备")
    start_dt, end_dt = _parse_range(start, end)
    report = await compute_outage_report(db, start=start_dt, end=end_dt, device_key=device)

    wb = Workbook()
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    center = Alignment(horizontal="center", vertical="center")

    # 明细 sheet
    ws = wb.active
    ws.title = "掉线明细"
    headers = ["设备名称", "异常开始时间", "异常结束时间", "持续时长", "是否仍在异常"]
    ws.append(headers)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center
    for row in report["rows"]:
        ws.append(
            [
                _device_label(row["device_name"], row["device_type"]),
                row["start"].replace("T", " ")[:19],
                row["end"].replace("T", " ")[:19],
                _format_duration_by_type(row["duration_seconds"], row["device_type"]),
                "是" if row["ongoing"] else "否",
            ]
        )
    for col, width in zip("ABCDE", (18, 21, 21, 14, 12)):
        ws.column_dimensions[col].width = width

    # 统计 sheet
    ws2 = wb.create_sheet("统计")
    ws2.append(["设备名称", "范围内次数", "范围内总时长", "近7天次数", "近7天总时长", "近30天次数", "近30天总时长"])
    for cell in ws2[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center
    for item in report["per_device"]:
        ws2.append(
            [
                _device_label(item["device_name"], item["device_type"]),
                item["range_count"],
                _format_duration(item["range_duration_seconds"]),
                item["week_count"],
                _format_duration(item["week_duration_seconds"]),
                item["month_count"],
                _format_duration(item["month_duration_seconds"]),
            ]
        )
    for col, width in zip("ABCDEFG", (18, 12, 16, 12, 16, 12, 16)):
        ws2.column_dimensions[col].width = width

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    # 中文文件名需 RFC 5987 编码（filename* ），并保留 ASCII 回退名。
    cn_name = f"监测平台设备状态运维-{cn_now_naive().strftime('%Y%m%d%H%M%S')}.xlsx"
    ascii_fallback = f"device_outages_{start_dt.strftime('%Y%m%d')}_{end_dt.strftime('%Y%m%d')}.xlsx"
    disposition = f"attachment; filename=\"{ascii_fallback}\"; filename*=UTF-8''{quote(cn_name)}"
    return Response(
        content=buffer.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": disposition},
    )
