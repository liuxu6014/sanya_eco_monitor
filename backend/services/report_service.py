"""Report generation service for the Sanya monitoring platform.

Provides aggregated summaries of insect, spore, rainfall, runoff, and water-quality data
over configurable date ranges, and can render those summaries as Excel
workbooks or standalone HTML documents.
"""

from __future__ import annotations

import io
from collections import defaultdict
from datetime import date, datetime, timedelta
from typing import Any

from jinja2 import Environment, BaseLoader
from openpyxl import Workbook
from openpyxl.styles import (
    Alignment,
    Font,
    PatternFill,
    Border,
    Side,
)
from sqlalchemy import desc, select
from sqlalchemy.ext.asyncio import AsyncSession

from config import settings
from models import (
    InsectRecord, SporeRecord,
    WaterQualityRecord, RainfallRecord, RunoffRecord
)
from services.guideline_metrics import build_guideline_metrics
from services.report_figures import build_figure_manifest
from services.water_quality_support import get_water_quality_records, resolve_water_quality_codes
from time_utils import cn_now_str


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _safe_avg(values: list[float]) -> float | None:
    """Return the arithmetic mean of *values*, or None when the list is empty."""
    clean = [v for v in values if v is not None]
    if not clean:
        return None
    return round(sum(clean) / len(clean), 2)


def _round_or_none(value: float | None, ndigits: int = 2) -> float | None:
    if value is None:
        return None
    return round(value, ndigits)


def _date_range_bounds(start_date: date, end_date: date) -> tuple[datetime, datetime]:
    """Convert date objects to datetime range covering full days (inclusive)."""
    start_dt = datetime(start_date.year, start_date.month, start_date.day, 0, 0, 0)
    end_dt = datetime(end_date.year, end_date.month, end_date.day, 23, 59, 59)
    return start_dt, end_dt


def _rounded_delta(value: float | None, ndigits: int = 2) -> float | int | None:
    if value is None:
        return None
    rounded = round(value, ndigits)
    if float(rounded).is_integer():
        return int(rounded)
    return rounded


def _change_rate(current: float | int | None, previous: float | int | None) -> float | None:
    if current is None or previous in (None, 0):
        return None
    return round((float(current) - float(previous)) / float(previous) * 100, 1)


def _trend_label(current: float | int | None, previous: float | int | None) -> str:
    if current is None or previous is None:
        return "暂无对比"
    diff = float(current) - float(previous)
    if abs(diff) < 1e-9:
        return "持平"
    return "上升" if diff > 0 else "下降"


def _build_period_delta(
    *,
    label: str,
    metric_label: str,
    unit: str,
    current_value: float | int | None,
    previous_value: float | int | None,
    ndigits: int = 2,
) -> dict[str, Any]:
    change_value = None
    if current_value is not None and previous_value is not None:
        change_value = _rounded_delta(float(current_value) - float(previous_value), ndigits=ndigits)

    return {
        "label": label,
        "metric_label": metric_label,
        "unit": unit,
        "current_value": current_value,
        "previous_value": previous_value,
        "change_value": change_value,
        "change_rate": _change_rate(current_value, previous_value),
        "trend": _trend_label(current_value, previous_value),
        "available": current_value is not None or previous_value is not None,
    }


def build_history_comparison_summary(
    *,
    current_period: dict[str, str],
    previous_period: dict[str, str],
    current: dict[str, Any],
    previous: dict[str, Any],
) -> dict[str, Any]:
    modules = {
        "insect": _build_period_delta(
            label="虫情测报",
            metric_label="周期内有效捕获昆虫",
            unit="只",
            current_value=(current.get("insect") or {}).get("total_count"),
            previous_value=(previous.get("insect") or {}).get("total_count"),
            ndigits=1,
        ),
        "spore": _build_period_delta(
            label="孢子监测",
            metric_label="周期内有效捕获孢子",
            unit="个",
            current_value=(current.get("spore") or {}).get("total_count"),
            previous_value=(previous.get("spore") or {}).get("total_count"),
            ndigits=1,
        ),
        "rain": _build_period_delta(
            label="雨量监测",
            metric_label="累计降雨量",
            unit="mm",
            current_value=(current.get("rain") or {}).get("total_rainfall"),
            previous_value=(previous.get("rain") or {}).get("total_rainfall"),
            ndigits=2,
        ),
        "runoff": _build_period_delta(
            label="地表径流监测",
            metric_label="累计径流量",
            unit="m3",
            current_value=(current.get("runoff") or {}).get("total_runoff"),
            previous_value=(previous.get("runoff") or {}).get("total_runoff"),
            ndigits=2,
        ),
    }

    water_metric_fields = (
        ("氨氮", "avg_nh3_n"),
        ("总磷", "avg_tp"),
        ("高锰酸盐指数", "avg_permanganate"),
        ("总氮", "avg_tn"),
    )
    water_metrics: list[dict[str, Any]] = []
    improved_count = 0
    degraded_count = 0
    for label, field_name in water_metric_fields:
        current_value = (current.get("water_quality") or {}).get(field_name)
        previous_value = (previous.get("water_quality") or {}).get(field_name)
        item = _build_period_delta(
            label=label,
            metric_label="周期平均值",
            unit="mg/L",
            current_value=current_value,
            previous_value=previous_value,
            ndigits=3,
        )
        water_metrics.append(item)
        if item["change_value"] is None:
            continue
        if float(item["change_value"]) < 0:
            improved_count += 1
        elif float(item["change_value"]) > 0:
            degraded_count += 1

    return {
        "comparison_basis": "本期与上一等长周期对比",
        "current_period": current_period,
        "previous_period": previous_period,
        "modules": modules,
        "water_quality": {
            "metric_label": "水质关键指标周期均值对比",
            "metrics": water_metrics,
            "improved_count": improved_count,
            "degraded_count": degraded_count,
        },
    }


def _display_value(value: Any, unit: str = "", fallback: str = "—") -> str:
    if value in (None, ""):
        return fallback
    return f"{value}{unit}"


def _daily_peak(rows: list[dict[str, Any]], key: str) -> dict[str, Any]:
    clean = [
        {
            "date": item.get("date"),
            "value": item.get(key) if item.get(key) is not None else item.get("count"),
        }
        for item in rows or []
    ]
    clean = [item for item in clean if item["date"] and item["value"] is not None]
    if not clean:
        return {"date": "—", "value": 0, "active_days": 0, "trend": "暂无趋势"}
    peak = max(clean, key=lambda item: item["value"])
    active_days = len([item for item in clean if (item.get("value") or 0) > 0])
    first = clean[0]["value"] or 0
    last = clean[-1]["value"] or 0
    if last > first:
        trend = "后段高于前段"
    elif last < first:
        trend = "后段低于前段"
    else:
        trend = "整体平稳"
    return {
        "date": peak["date"],
        "value": peak["value"],
        "active_days": active_days,
        "trend": trend,
    }


def build_special_analysis_sections(summary: dict[str, Any]) -> list[dict[str, Any]]:
    """Build fixed deep analysis content for HTML/DOCX reports, excluding spore text."""
    ins = summary.get("insect", {}) or {}
    rn = summary.get("rain", {}) or {}
    ro = summary.get("runoff", {}) or {}
    wq = summary.get("water_quality", {}) or {}
    gm = summary.get("guideline_metrics", {}) or {}
    history = (summary.get("history_comparison", {}) or {}).get("modules", {}) or {}
    water_history = (summary.get("history_comparison", {}) or {}).get("water_quality", {}) or {}
    runoff_guideline = gm.get("runoff_erosion", {}) or {}
    water_guideline = gm.get("water_quality", {}) or {}
    pest_guideline = gm.get("pest_management", {}) or {}
    warning_analysis = gm.get("warning_analysis", {}) or {}

    insect_peak = _daily_peak(ins.get("daily") or [], "count")
    rain_peak = _daily_peak(rn.get("daily") or [], "rainfall")
    top_species = (ins.get("top_species") or [["暂无", 0]])[0]
    warning_items = {
        item.get("key"): item
        for item in (warning_analysis.get("indicator_warnings") or [])
    }

    runoff_devices = ro.get("by_device") or {}
    device_rows = list(runoff_devices.values())
    max_sand = max(
        device_rows,
        key=lambda item: item.get("avg_sand_content") or 0,
        default={},
    )
    max_runoff = max(
        device_rows,
        key=lambda item: item.get("total_runoff") or 0,
        default={},
    )

    water_metrics = water_guideline.get("metrics") or []
    weakest_water = max(
        water_metrics,
        key=lambda item: item.get("latest_value") or item.get("recent_avg") or 0,
        default={},
    )

    insect_history = history.get("insect") or {}
    rain_history = history.get("rain") or {}
    runoff_history = history.get("runoff") or {}

    return [
        {
            "title": "虫情深度专项分析",
            "badge": pest_guideline.get("risk_level", "风险研判"),
            "facts": [
                f"累计捕获 {_display_value(ins.get('total_count'), '只', '0只')}",
                f"记录 {_display_value(ins.get('records_count'), '条', '0条')}",
                f"优势虫种 {top_species[0]}",
                f"峰值 {insect_peak['date']} / {_display_value(insect_peak['value'], '只')}",
            ],
            "paragraphs": [
                (
                    f"本期虫情监测形成 {ins.get('records_count', 0)} 条有效记录，累计捕获 "
                    f"{ins.get('total_count', 0)} 只。优势虫种为 {top_species[0]}，数量为 {top_species[1] if len(top_species) > 1 else 0} 只；"
                    f"日尺度峰值出现在 {insect_peak['date']}，单日捕获 {insect_peak['value']} 只，趋势表现为{insect_peak['trend']}。"
                ),
                (
                    f"历史同口径对比显示，虫情捕获较上一周期为{insect_history.get('trend', '暂无对比')}，"
                    f"变化率为 {_display_value(insect_history.get('change_rate'), '%')}。该项应作为生物多样性变化与害虫风险的双重信号，"
                    "既要关注数量异常升高，也要关注高危害虫种在短时段内连续出现。"
                ),
            ],
            "actions": [
                "将优势虫种纳入重点巡查清单，复核诱捕点周边作物叶片、嫩梢和果实危害情况。",
                "当虫情峰值与高温高湿天气叠加时，提高灯诱、性诱和田间样方调查频次。",
                "处置上优先采用物理诱控、生物防治和低毒精准药剂，避免无差别大范围用药。",
            ],
        },
        {
            "title": "雨情深度专项分析",
            "badge": warning_items.get("rainfall_peak", {}).get("level", "雨量研判"),
            "facts": [
                f"累计降雨 {_display_value(rn.get('total_rainfall'), 'mm', '0mm')}",
                f"记录 {_display_value(rn.get('records_count'), '条', '0条')}",
                f"雨日 {_display_value(rain_peak['active_days'], '天')}",
                f"峰值 {rain_peak['date']} / {_display_value(rain_peak['value'], 'mm')}",
            ],
            "paragraphs": [
                (
                    f"本期雨量监测累计降雨 {rn.get('total_rainfall', 0)} mm，形成 {rn.get('records_count', 0)} 条记录，"
                    f"有降雨记录的天数为 {rain_peak['active_days']} 天。日尺度峰值出现在 {rain_peak['date']}，"
                    f"峰值雨量为 {rain_peak['value']} mm，趋势表现为{rain_peak['trend']}。"
                ),
                (
                    f"历史同口径对比显示，累计雨量较上一周期为{rain_history.get('trend', '暂无对比')}，"
                    f"变化率为 {_display_value(rain_history.get('change_rate'), '%')}。雨情是径流、含沙量和面源污染迁移的触发因子，"
                    "强降雨后应重点观察坡面裸露区、排水出口和低洼汇流点。"
                ),
            ],
            "actions": [
                "雨前检查雨量计、径流沟渠、采样容器和通信供电，保证关键降雨过程数据连续。",
                "单日雨量超过预警阈值时，联动查看径流量、流速、含沙量和水质指标。",
                "雨后优先安排低洼地块、坡面裸露区域和排水出口巡查，记录冲刷与淤积点位。",
            ],
        },
        {
            "title": "水土流失与径流深度专项分析",
            "badge": "径流含沙联动",
            "facts": [
                f"累计径流 {_display_value(ro.get('total_runoff'), 'm3', '0m3')}",
                f"记录 {_display_value(ro.get('records_count'), '条', '0条')}",
                f"监测点 {_display_value(ro.get('device_count'), '个', '0个')}",
                f"减蚀率 {_display_value(runoff_guideline.get('estimated_reduction_rate'), '%')}",
            ],
            "paragraphs": [
                (
                    f"本期地表径流监测覆盖 {ro.get('device_count', 0)} 个监测点，形成 {ro.get('records_count', 0)} 条记录，"
                    f"累计径流量为 {ro.get('total_runoff', 0)} m3。径流量最高的监测点为 {max_runoff.get('name', '—')}，"
                    f"累计径流量为 {_display_value(max_runoff.get('total_runoff'), 'm3')}；平均含沙量较高的监测点为 {max_sand.get('name', '—')}，"
                    f"平均含沙量为 {_display_value(max_sand.get('avg_sand_content'))}。"
                ),
                (
                    f"历史同口径对比显示，累计径流量较上一周期为{runoff_history.get('trend', '暂无对比')}，"
                    f"变化率为 {_display_value(runoff_history.get('change_rate'), '%')}。监测型估算减蚀率为 "
                    f"{_display_value(runoff_guideline.get('estimated_reduction_rate'), '%')}，可用于判断近自然化改造对坡面拦截、地表覆盖和水源涵养的阶段性效果。"
                ),
            ],
            "actions": [
                "对径流量和含沙量同步偏高的监测点开展现场复核，重点查看裸露坡面、汇流沟和排水出口。",
                "在强降雨后增加水土保持设施巡查频次，记录冲刷沟、淤积带和植被覆盖缺口。",
                "对比次生林参照点与经营林地监测点差异，持续优化覆盖层、缓冲带和排水组织。",
            ],
        },
        {
            "title": "面源水质污染深度专项分析",
            "badge": _display_value(water_guideline.get("composite_reduction_rate"), "%", "水质研判"),
            "facts": [
                f"记录 {_display_value(wq.get('records_count'), '条', '0条')}",
                f"氨氮 {_display_value(wq.get('avg_nh3_n'), 'mg/L')}",
                f"总磷 {_display_value(wq.get('avg_tp'), 'mg/L')}",
                f"总氮 {_display_value(wq.get('avg_tn'), 'mg/L')}",
            ],
            "paragraphs": [
                (
                    f"本期水质监测形成 {wq.get('records_count', 0)} 条记录，平均氨氮为 {_display_value(wq.get('avg_nh3_n'), 'mg/L')}，"
                    f"平均总磷为 {_display_value(wq.get('avg_tp'), 'mg/L')}，平均高锰酸盐指数为 {_display_value(wq.get('avg_permanganate'), 'mg/L')}，"
                    f"平均总氮为 {_display_value(wq.get('avg_tn'), 'mg/L')}。当前相对突出的指标为 {weakest_water.get('label', '—')}，"
                    f"最新值为 {_display_value(weakest_water.get('latest_value'), weakest_water.get('unit', ''))}。"
                ),
                (
                    f"农业面源污染综合削减率为 {_display_value(water_guideline.get('composite_reduction_rate'), '%')}。"
                    f"水质关键指标中，较上一周期改善 {water_history.get('improved_count', 0)} 项、升高 {water_history.get('degraded_count', 0)} 项。"
                    "该部分应与降雨和径流过程联合解释：降雨驱动径流形成，径流携带氮磷和有机污染物进入汇水通道，生态缓冲带和林下覆盖则决定污染拦截效率。"
                ),
            ],
            "actions": [
                "强降雨后优先复核氨氮、总磷、总氮和高锰酸盐指数，关注短时冲刷造成的浓度波动。",
                "对水质偏高时段追溯对应降雨、径流和含沙量过程，判断是否存在面源输入集中释放。",
                "持续优化农田排水口、生态沟渠和林缘缓冲带，提升氮磷拦截和颗粒物沉降能力。",
            ],
        },
    ]


# ---------------------------------------------------------------------------
# ReportService
# ---------------------------------------------------------------------------

class ReportService:
    """Aggregate monitoring data and produce reports."""

    async def _fallback_capture_images(
        self,
        db: AsyncSession,
        model,
        end_dt: datetime,
        *,
        limit: int = 3,
    ) -> list[dict]:
        result = await db.execute(
            select(model)
            .where(
                model.collection_time <= end_dt,
                model.image_url.is_not(None),
                model.image_url != "",
            )
            .order_by(desc(model.collection_time))
            .limit(limit)
        )
        records = list(result.scalars().all())
        records.reverse()
        return [
            {
                "time": record.collection_time.strftime("%Y-%m-%d %H:%M"),
                "device_code": record.device_code,
                "url": record.image_url,
            }
            for record in records
            if record.image_url
        ]

    # ------------------------------------------------------------------
    # Public summary entry points
    # ------------------------------------------------------------------

    async def get_week_summary(
        self,
        db: AsyncSession,
        end_date: date | None = None,
    ) -> dict[str, Any]:
        """Return an aggregated summary for the 7 days ending on *end_date*."""
        if end_date is None:
            end_date = date.today()
        start_date = end_date - timedelta(days=6)
        return await self._build_summary(db, start_date, end_date)

    async def get_month_summary(
        self,
        db: AsyncSession,
        end_date: date | None = None,
    ) -> dict[str, Any]:
        """Return an aggregated summary for the 30 days ending on *end_date*."""
        if end_date is None:
            end_date = date.today()
        start_date = end_date - timedelta(days=29)
        return await self._build_summary(db, start_date, end_date)

    async def get_custom_summary(
        self,
        db: AsyncSession,
        start_date: date,
        end_date: date,
    ) -> dict[str, Any]:
        """Return an aggregated summary for the given date range (inclusive)."""
        return await self._build_summary(db, start_date, end_date)

    # ------------------------------------------------------------------
    # Core aggregation
    # ------------------------------------------------------------------

    async def _build_summary(
        self,
        db: AsyncSession,
        start_date: date,
        end_date: date,
    ) -> dict[str, Any]:
        start_dt, end_dt = _date_range_bounds(start_date, end_date)
        span_days = (end_date - start_date).days + 1
        previous_end_date = start_date - timedelta(days=1)
        previous_start_date = previous_end_date - timedelta(days=span_days - 1)
        previous_start_dt, previous_end_dt = _date_range_bounds(previous_start_date, previous_end_date)

        insect = await self._aggregate_insect(db, start_dt, end_dt)
        spore = await self._aggregate_spore(db, start_dt, end_dt)
        water_quality = await self._aggregate_water_quality(db, start_dt, end_dt)
        rain = await self._aggregate_rainfall(db, start_dt, end_dt)
        runoff = await self._aggregate_runoff(db, start_dt, end_dt)
        previous_insect = await self._aggregate_insect(db, previous_start_dt, previous_end_dt)
        previous_spore = await self._aggregate_spore(db, previous_start_dt, previous_end_dt)
        previous_water_quality = await self._aggregate_water_quality(db, previous_start_dt, previous_end_dt)
        previous_rain = await self._aggregate_rainfall(db, previous_start_dt, previous_end_dt)
        previous_runoff = await self._aggregate_runoff(db, previous_start_dt, previous_end_dt)
        guideline_metrics = await build_guideline_metrics(db)
        current_period = {
            "start": start_date.strftime("%Y-%m-%d"),
            "end": end_date.strftime("%Y-%m-%d"),
        }
        previous_period = {
            "start": previous_start_date.strftime("%Y-%m-%d"),
            "end": previous_end_date.strftime("%Y-%m-%d"),
        }
        history_comparison = build_history_comparison_summary(
            current_period=current_period,
            previous_period=previous_period,
            current={
                "insect": insect,
                "spore": spore,
                "water_quality": water_quality,
                "rain": rain,
                "runoff": runoff,
            },
            previous={
                "insect": previous_insect,
                "spore": previous_spore,
                "water_quality": previous_water_quality,
                "rain": previous_rain,
                "runoff": previous_runoff,
            },
        )

        return {
            "period": current_period,
            "insect": insect,
            "spore": spore,
            "water_quality": water_quality,
            "rain": rain,
            "runoff": runoff,
            "history_comparison": history_comparison,
            "guideline_metrics": guideline_metrics,
            "weather_support": guideline_metrics.get("weather_support"),
        }

    # ------------------------------------------------------------------
    # Per-model aggregators
    # ------------------------------------------------------------------

    async def _aggregate_insect(
        self,
        db: AsyncSession,
        start_dt: datetime,
        end_dt: datetime,
    ) -> dict[str, Any]:
        result = await db.execute(
            select(InsectRecord).where(
                InsectRecord.collection_time >= start_dt,
                InsectRecord.collection_time <= end_dt,
            ).order_by(InsectRecord.collection_time)
        )
        records = result.scalars().all()

        total_count = sum(r.total_count for r in records)

        # Aggregate species counts across all records
        species_totals: dict[str, int] = defaultdict(int)
        for r in records:
            if r.species_data:
                for name, cnt in r.species_data.items():
                    species_totals[name] += int(cnt)

        top_species = sorted(
            species_totals.items(), key=lambda x: x[1], reverse=True
        )[:10]

        # Daily aggregation
        daily: dict[str, int] = defaultdict(int)
        for r in records:
            day_key = r.collection_time.strftime("%Y-%m-%d")
            daily[day_key] += r.total_count

        daily_list = [
            {"date": d, "count": c}
            for d, c in sorted(daily.items())
        ]

        # Collect capture images grouped by device
        capture_images: list[dict] = [
            {
                "time": r.collection_time.strftime("%Y-%m-%d %H:%M"),
                "device_code": r.device_code,
                "url": r.image_url,
            }
            for r in records
            if r.image_url
        ]
        if not capture_images:
            capture_images = await self._fallback_capture_images(
                db,
                InsectRecord,
                end_dt,
            )

        return {
            "total_count": total_count,
            "records_count": len(records),
            "top_species": [list(item) for item in top_species],
            "daily": daily_list,
            "capture_images": capture_images,
        }

    async def _aggregate_spore(
        self,
        db: AsyncSession,
        start_dt: datetime,
        end_dt: datetime,
    ) -> dict[str, Any]:
        result = await db.execute(
            select(SporeRecord).where(
                SporeRecord.collection_time >= start_dt,
                SporeRecord.collection_time <= end_dt,
            ).order_by(SporeRecord.collection_time)
        )
        records = result.scalars().all()

        total_count = sum(r.total_count for r in records)

        daily: dict[str, int] = defaultdict(int)
        for r in records:
            day_key = r.collection_time.strftime("%Y-%m-%d")
            daily[day_key] += r.total_count

        daily_list = [
            {"date": d, "count": c}
            for d, c in sorted(daily.items())
        ]

        # Collect capture images
        capture_images: list[dict] = [
            {
                "time": r.collection_time.strftime("%Y-%m-%d %H:%M"),
                "device_code": r.device_code,
                "url": r.image_url,
            }
            for r in records
            if r.image_url
        ]
        if not capture_images:
            capture_images = await self._fallback_capture_images(
                db,
                SporeRecord,
                end_dt,
            )

        return {
            "total_count": total_count,
            "records_count": len(records),
            "daily": daily_list,
            "capture_images": capture_images,
        }

    async def _aggregate_water_quality(
        self,
        db: AsyncSession,
        start_dt: datetime,
        end_dt: datetime,
    ) -> dict[str, Any]:
        configured_water_code = settings.WATER_QUALITY_CODE.strip() or "16133028"
        water_codes = await resolve_water_quality_codes(
            db,
            preferred_code=configured_water_code,
            start_dt=start_dt,
            end_dt=end_dt,
        )
        records = await get_water_quality_records(
            db,
            start_dt=start_dt,
            end_dt=end_dt,
            codes=water_codes,
        )
        if not records:
            return {"records_count": 0}

        nh3s = [r.ammonia_nitrogen for r in records if r.ammonia_nitrogen is not None]
        tps = [r.total_phosphorus for r in records if r.total_phosphorus is not None]
        tns = [r.total_nitrogen for r in records if r.total_nitrogen is not None]
        permanganates = [r.permanganate_index for r in records if r.permanganate_index is not None]

        return {
            "records_count": len(records),
            "device_code": records[-1].device_code,
            "configured_device_code": configured_water_code,
            "avg_nh3_n": _safe_avg(nh3s),
            "avg_tp": _safe_avg(tps),
            "avg_tn": _safe_avg(tns),
            "avg_permanganate": _safe_avg(permanganates),
        }

    async def _aggregate_rainfall(
        self,
        db: AsyncSession,
        start_dt: datetime,
        end_dt: datetime,
    ) -> dict[str, Any]:
        result = await db.execute(
            select(RainfallRecord).where(
                RainfallRecord.collection_time >= start_dt,
                RainfallRecord.collection_time <= end_dt,
            )
        )
        records = result.scalars().all()
        if not records:
            return {"records_count": 0, "total_rainfall": 0.0, "daily": []}

        daily_totals: dict[str, float] = defaultdict(float)
        total_rainfall = 0.0
        for record in records:
            rainfall_value = record.hourly_rainfall
            if rainfall_value is None:
                rainfall_value = record.rainfall
            if rainfall_value is None:
                continue
            total_rainfall += rainfall_value
            daily_totals[record.collection_time.strftime("%Y-%m-%d")] += rainfall_value

        return {
            "records_count": len(records),
            "total_rainfall": round(total_rainfall, 2),
            "daily": [
                {"date": day, "rainfall": round(value, 2)}
                for day, value in sorted(daily_totals.items())
            ],
        }

    async def _aggregate_runoff(
        self,
        db: AsyncSession,
        start_dt: datetime,
        end_dt: datetime,
    ) -> dict[str, Any]:
        DEVICE_NAMES = {
            '16132920': '杧果林1监测点',
            '16132921': '橡胶林1监测点',
            '16132922': '次生林监测点',
            '16132923': '杧果林2监测点',
            '16132924': '橡胶林2监测点',
            '16132925': '槟榔林监测点',
        }

        result = await db.execute(
            select(RunoffRecord).where(
                RunoffRecord.collection_time >= start_dt,
                RunoffRecord.collection_time <= end_dt,
            ).order_by(RunoffRecord.collection_time)
        )
        records = result.scalars().all()
        if not records:
            return {"records_count": 0, "by_device": {}, "total_runoff": 0.0}

        # Per-device aggregation
        by_device: dict[str, dict] = defaultdict(lambda: {
            "flow_speeds": [], "flow_rates": [], "total_flows": [],
            "water_levels": [], "sand_contents": [], "liquid_pressures": [],
            "runoffs": [], "rainfalls": [], "count": 0
        })

        for r in records:
            d = by_device[r.device_code]
            d["count"] += 1
            for field, lst in [
                (r.flow_speed, "flow_speeds"), (r.flow_rate, "flow_rates"),
                (r.total_flow, "total_flows"), (r.water_level, "water_levels"),
                (r.sand_content, "sand_contents"), (r.liquid_pressure, "liquid_pressures"),
                (r.runoff, "runoffs"), (r.rainfall, "rainfalls"),
            ]:
                if field is not None:
                    d[lst].append(field)

        device_summary = {}
        for code, d in by_device.items():
            device_summary[code] = {
                "name": DEVICE_NAMES.get(code, code),
                "records_count": d["count"],
                "avg_flow_speed":     _safe_avg(d["flow_speeds"]),
                "max_flow_speed":     _round_or_none(max(d["flow_speeds"])) if d["flow_speeds"] else None,
                "avg_flow_rate":      _safe_avg(d["flow_rates"]),
                "max_flow_rate":      _round_or_none(max(d["flow_rates"])) if d["flow_rates"] else None,
                "total_flow_latest":  _round_or_none(d["total_flows"][-1]) if d["total_flows"] else None,
                "avg_water_level":    _safe_avg(d["water_levels"]),
                "max_water_level":    _round_or_none(max(d["water_levels"])) if d["water_levels"] else None,
                "avg_sand_content":   _safe_avg(d["sand_contents"]),
                "avg_liquid_pressure":_safe_avg(d["liquid_pressures"]),
                "total_runoff":       _round_or_none(sum(d["runoffs"])) if d["runoffs"] else 0.0,
                "total_rainfall":     _round_or_none(sum(d["rainfalls"])) if d["rainfalls"] else 0.0,
            }

        # Overall aggregated values across all devices
        all_flows = [r.flow_rate for r in records if r.flow_rate is not None]
        all_levels = [r.water_level for r in records if r.water_level is not None]

        return {
            "records_count": len(records),
            "device_count": len(by_device),
            "avg_flow_rate": _safe_avg(all_flows),
            "max_flow_rate": _round_or_none(max(all_flows)) if all_flows else None,
            "avg_water_level": _safe_avg(all_levels),
            "total_runoff": _round_or_none(
                sum(item.get("total_runoff", 0.0) for item in device_summary.values())
            )
            or 0.0,
            "by_device": device_summary,
        }

    # ------------------------------------------------------------------
    # Excel export
    # ------------------------------------------------------------------

    async def generate_excel(
        self,
        db: AsyncSession,
        start_date: date,
        end_date: date,
    ) -> bytes:
        """Create a multi-sheet Excel workbook and return the raw bytes."""
        summary = await self._build_summary(db, start_date, end_date)
        start_dt, end_dt = _date_range_bounds(start_date, end_date)

        insect_res = await db.execute(
            select(InsectRecord).where(
                InsectRecord.collection_time >= start_dt,
                InsectRecord.collection_time <= end_dt,
            ).order_by(InsectRecord.collection_time)
        )
        insect_records = insect_res.scalars().all()

        wb = Workbook()

        # Remove default sheet
        default_sheet = wb.active
        wb.remove(default_sheet)

        self._write_summary_sheet(wb, summary)
        self._write_insect_sheet(wb, insect_records)

        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()

    # ------------------------------------------------------------------
    # Excel sheet writers
    # ------------------------------------------------------------------

    @staticmethod
    def _header_fill() -> PatternFill:
        return PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")

    @staticmethod
    def _sub_header_fill() -> PatternFill:
        return PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")

    @staticmethod
    def _header_font() -> Font:
        return Font(name="微软雅黑", bold=True, color="FFFFFF", size=11)

    @staticmethod
    def _title_font() -> Font:
        return Font(name="微软雅黑", bold=True, color="FFFFFF", size=13)

    @staticmethod
    def _label_font() -> Font:
        return Font(name="微软雅黑", bold=True, color="1F4E79", size=10)

    @staticmethod
    def _value_font() -> Font:
        return Font(name="微软雅黑", size=10)

    @staticmethod
    def _thin_border() -> Border:
        side = Side(style="thin", color="BFBFBF")
        return Border(left=side, right=side, top=side, bottom=side)

    @staticmethod
    def _center_align() -> Alignment:
        return Alignment(horizontal="center", vertical="center", wrap_text=True)

    def _write_title_row(
        self,
        ws,
        title: str,
        col_span: int,
        row: int = 1,
    ) -> None:
        ws.cell(row=row, column=1, value=title).font = self._title_font()
        ws.cell(row=row, column=1).fill = PatternFill(
            start_color="1F4E79", end_color="1F4E79", fill_type="solid"
        )
        ws.cell(row=row, column=1).alignment = self._center_align()
        ws.merge_cells(
            start_row=row, start_column=1, end_row=row, end_column=col_span
        )

    def _apply_header_row(self, ws, headers: list[str], row: int) -> None:
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=row, column=col_idx, value=header)
            cell.font = self._header_font()
            cell.fill = self._sub_header_fill()
            cell.alignment = self._center_align()
            cell.border = self._thin_border()

    def _write_summary_sheet(self, wb: Workbook, summary: dict) -> None:
        ws = wb.create_sheet("综合汇总")
        ws.column_dimensions["A"].width = 22
        ws.column_dimensions["B"].width = 18
        ws.column_dimensions["C"].width = 22
        ws.column_dimensions["D"].width = 18

        period = summary["period"]
        title = (
            f"三亚市天涯区橡胶林近自然化改造和农田提升监测平台  "
            f"监测报告 {period['start']} 至 {period['end']}"
        )
        self._write_title_row(ws, title, 4, row=1)
        ws.row_dimensions[1].height = 28

        row = 3
        ws.cell(row=row, column=1, value="【虫情测报】").font = self._label_font()
        ws.cell(row=row, column=1).fill = PatternFill(
            start_color="DEEAF1", end_color="DEEAF1", fill_type="solid"
        )
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
        row += 1

        ins = summary["insect"]
        insect_rows = [
            ("捕获总数 (只)", ins["total_count"], "记录条数", ins["records_count"]),
            ("主要虫种数", len(ins.get("top_species") or []), "实拍图数量", len(ins.get("capture_images") or [])),
        ]
        for label_a, val_a, label_b, val_b in insect_rows:
            ws.cell(row=row, column=1, value=label_a).font = self._label_font()
            ws.cell(row=row, column=2, value=val_a if val_a is not None else "—").font = self._value_font()
            ws.cell(row=row, column=3, value=label_b).font = self._label_font()
            ws.cell(row=row, column=4, value=val_b if val_b is not None else "—").font = self._value_font()
            for col in range(1, 5):
                ws.cell(row=row, column=col).border = self._thin_border()
                ws.cell(row=row, column=col).alignment = self._center_align()
            row += 1

        row += 1
        ws.cell(row=row, column=1, value="【孢子捕捉】").font = self._label_font()
        ws.cell(row=row, column=1).fill = PatternFill(
            start_color="DEEAF1", end_color="DEEAF1", fill_type="solid"
        )
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
        row += 1

        sp = summary["spore"]
        ws.cell(row=row, column=1, value="孢子总数 (个)").font = self._label_font()
        ws.cell(row=row, column=2, value=sp["total_count"]).font = self._value_font()
        ws.cell(row=row, column=3, value="记录条数").font = self._label_font()
        ws.cell(row=row, column=4, value=sp["records_count"]).font = self._value_font()
        for col in range(1, 5):
            ws.cell(row=row, column=col).border = self._thin_border()
            ws.cell(row=row, column=col).alignment = self._center_align()
        row += 1

        row += 1
        ws.cell(row=row, column=1, value="【雨量与径流】").font = self._label_font()
        ws.cell(row=row, column=1).fill = PatternFill(
            start_color="DEEAF1", end_color="DEEAF1", fill_type="solid"
        )
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
        row += 1

        rain = summary["rain"]
        runoff = summary["runoff"]
        hydro_rows = [
            ("雨量记录条数", rain["records_count"], "累计降雨量 (mm)", rain["total_rainfall"]),
            ("径流记录条数", runoff["records_count"], "监测设备数", runoff["device_count"]),
            ("平均流量", runoff["avg_flow_rate"], "平均水位", runoff["avg_water_level"]),
        ]
        for label_a, val_a, label_b, val_b in hydro_rows:
            ws.cell(row=row, column=1, value=label_a).font = self._label_font()
            ws.cell(row=row, column=2, value=val_a if val_a is not None else "—").font = self._value_font()
            ws.cell(row=row, column=3, value=label_b).font = self._label_font()
            ws.cell(row=row, column=4, value=val_b if val_b is not None else "—").font = self._value_font()
            for col in range(1, 5):
                ws.cell(row=row, column=col).border = self._thin_border()
                ws.cell(row=row, column=col).alignment = self._center_align()
            row += 1

        row += 1
        ws.cell(row=row, column=1, value="【水质监测】").font = self._label_font()
        ws.cell(row=row, column=1).fill = PatternFill(
            start_color="DEEAF1", end_color="DEEAF1", fill_type="solid"
        )
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
        row += 1

        wq = summary["water_quality"]
        water_rows = [
            ("水质记录条数", wq["records_count"], "平均氨氮", wq["avg_nh3_n"]),
            ("平均总磷", wq["avg_tp"], "平均高猛酸盐", wq["avg_permanganate"]),
            ("平均总氮", wq["avg_tn"], "监测设备", wq.get("device_code") or settings.WATER_QUALITY_CODE),
        ]
        for label_a, val_a, label_b, val_b in water_rows:
            ws.cell(row=row, column=1, value=label_a).font = self._label_font()
            ws.cell(row=row, column=2, value=val_a if val_a is not None else "—").font = self._value_font()
            ws.cell(row=row, column=3, value=label_b).font = self._label_font()
            ws.cell(row=row, column=4, value=val_b if val_b is not None else "—").font = self._value_font()
            for col in range(1, 5):
                ws.cell(row=row, column=col).border = self._thin_border()
                ws.cell(row=row, column=col).alignment = self._center_align()
            row += 1

        if ins["top_species"]:
            row += 1
            ws.cell(row=row, column=1, value="主要虫种统计").font = self._label_font()
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
            ws.cell(row=row, column=1).alignment = self._center_align()
            row += 1
            for name, count in ins["top_species"]:
                ws.cell(row=row, column=1, value=name).font = self._value_font()
                ws.cell(row=row, column=2, value=count).font = self._value_font()
                ws.cell(row=row, column=1).border = self._thin_border()
                ws.cell(row=row, column=2).border = self._thin_border()
                ws.cell(row=row, column=1).alignment = self._center_align()
                ws.cell(row=row, column=2).alignment = self._center_align()
                row += 1

    def _write_insect_sheet(self, wb: Workbook, records: list[InsectRecord]) -> None:
        ws = wb.create_sheet("虫情数据")
        headers = ["采集时间", "设备编号", "捕获数量 (只)", "主要虫种"]
        for col_idx, w in enumerate([22, 24, 16, 40], start=1):
            ws.column_dimensions[
                ws.cell(row=1, column=col_idx).column_letter
            ].width = w

        self._write_title_row(ws, "虫情测报详细数据", len(headers), row=1)
        ws.row_dimensions[1].height = 24
        self._apply_header_row(ws, headers, row=2)

        for row_idx, r in enumerate(records, start=3):
            top = sorted(
                (r.species_data or {}).items(), key=lambda x: x[1], reverse=True
            )[:3]
            species_str = "  ".join(f"{n}:{c}" for n, c in top) if top else "—"
            row_data = [
                r.collection_time.strftime("%Y-%m-%d %H:%M"),
                r.device_code,
                r.total_count,
                species_str,
            ]
            for col_idx, val in enumerate(row_data, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=val)
                cell.font = self._value_font()
                cell.border = self._thin_border()
                cell.alignment = Alignment(horizontal="center", vertical="center")
            if row_idx % 2 == 0:
                for col_idx in range(1, len(headers) + 1):
                    ws.cell(row=row_idx, column=col_idx).fill = PatternFill(
                        start_color="EBF3FB", end_color="EBF3FB", fill_type="solid"
                    )

    # ------------------------------------------------------------------
    # HTML report
    # ------------------------------------------------------------------

    @staticmethod
    def _render_ai_html(text: str) -> str:
        """Convert AI markdown output (## / ### / **1.1 bold** / **bold**) to HTML."""
        import re
        # Pattern: line is ONLY **x.x 小节名** (subsection heading style)
        _subsection_re = re.compile(r"^\*\*(\d+\.\d+\s+[^*]+)\*\*$")
        lines = text.split("\n")
        html_lines: list[str] = []
        in_para = False
        for line in lines:
            stripped = line.strip()
            if not stripped:
                if in_para:
                    html_lines.append("</p>")
                    in_para = False
                continue
            # Chapter heading  ## 一、xxx
            if stripped.startswith("## "):
                if in_para:
                    html_lines.append("</p>")
                    in_para = False
                content = stripped[3:].strip()
                content = re.sub(r"\*\*(.+?)\*\*", r"<strong>\1</strong>", content)
                html_lines.append(f'<h2 class="ai-h2">{content}</h2>')
            # Sub-section  ### xxx
            elif stripped.startswith("### "):
                if in_para:
                    html_lines.append("</p>")
                    in_para = False
                content = stripped[4:].strip()
                content = re.sub(r"\*\*(.+?)\*\*", r"<strong>\1</strong>", content)
                html_lines.append(f'<h3 class="ai-h3">{content}</h3>')
            # Sub-section heading style: **1.1 小节名**  (entire line is bold subsection)
            elif _subsection_re.match(stripped):
                if in_para:
                    html_lines.append("</p>")
                    in_para = False
                content = _subsection_re.match(stripped).group(1).strip()
                html_lines.append(f'<h3 class="ai-h3">{content}</h3>')
            # Bullet
            elif stripped.startswith(("- ", "• ", "* ")):
                if in_para:
                    html_lines.append("</p>")
                    in_para = False
                content = stripped[2:].strip()
                content = re.sub(r"\*\*(.+?)\*\*", r"<strong>\1</strong>", content)
                html_lines.append(f'<li>{content}</li>')
            else:
                content = re.sub(r"\*\*(.+?)\*\*", r"<strong>\1</strong>", stripped)
                if not in_para:
                    html_lines.append("<p>")
                    in_para = True
                else:
                    html_lines.append(" ")
                html_lines.append(content)
        if in_para:
            html_lines.append("</p>")
        return "\n".join(html_lines)

    @staticmethod
    def generate_html_report(
        summary_dict: dict,
        ai_analysis: str | None = None,
        charts: dict | None = None,
        ai_images: dict | None = None,
        figure_manifest: list[dict[str, Any]] | None = None,
    ) -> str:
        """Render *summary_dict* as a professional self-contained HTML report.

        Args:
            summary_dict: Aggregated data from ReportService.
            ai_analysis:  Pre-generated AI analysis text (markdown).
            charts:       Dict of chart_name → base64 PNG string (from chart_service).
            ai_images:    Dict from gemini_image_service.generate_report_images().
        """
        charts = charts or {}
        ai_images = ai_images or {}
        figure_manifest = figure_manifest or build_figure_manifest(summary_dict, charts, ai_images)

        period = summary_dict.get("period", {})
        start_str = period.get("start", "")
        end_str = period.get("end", "")

        try:
            s = datetime.strptime(start_str, "%Y-%m-%d").date()
            e = datetime.strptime(end_str, "%Y-%m-%d").date()
            span = (e - s).days + 1
            report_type = "周报" if span <= 7 else ("月报" if span <= 31 else "自定义报告")
        except (ValueError, TypeError):
            report_type = "自定义报告"

        ai_html = ReportService._render_ai_html(ai_analysis) if ai_analysis else ""

        # Gemini AI images
        cover_img_src = ai_images.get("cover")

        def _v(val, unit: str = "", fallback: str = "—") -> str:
            if val is None:
                return fallback
            return f"{val}{unit}"

        def _pct(val: float | None) -> str:
            return f"{val}%" if val is not None else "—"

        rn = summary_dict.get("rain", {})
        ro = summary_dict.get("runoff", {})
        wq = summary_dict.get("water_quality", {})
        ins = summary_dict.get("insect", {})
        sp = summary_dict.get("spore", {})
        gm = summary_dict.get("guideline_metrics", {}) or {}
        history_comparison = summary_dict.get("history_comparison", {}) or {}
        weather = summary_dict.get("weather_support", {}) or gm.get("weather_support", {}) or {}
        methodology = gm.get("methodology", {}) or {}
        runoff_guideline = gm.get("runoff_erosion", {}) or {}
        water_guideline = gm.get("water_quality", {}) or {}
        pest_guideline = gm.get("pest_management", {}) or {}
        warning_analysis = gm.get("warning_analysis", {}) or {}
        water_source_support = gm.get("water_source_support", {}) or {}
        implementation_matrix = gm.get("implementation_matrix", {}) or {}

        generated_at = cn_now_str()

        # Cover background style (Gemini-generated image)
        cover_style = (
            f'style="background-image:linear-gradient(160deg,rgba(13,33,55,0.82) 0%,'
            f'rgba(10,61,98,0.72) 55%,rgba(16,78,139,0.80) 100%),'
            f'url(\'{cover_img_src}\');background-size:cover;background-position:center;"'
            if cover_img_src else ""
        )

        def _is_pest_figure(item: dict[str, Any]) -> bool:
            return item["html_id"].startswith("fig-pest-")

        def _render_standard_figure(item: dict[str, Any]) -> str:
            tag_html = (
                f'&nbsp;<span class="gemini-tag">{item["tag"]}</span>'
                if item.get("tag")
                else ""
            )
            class_name = "chart-figure ai-img-figure" if item["source"] == "ai" else "chart-figure"
            return (
                f'<figure class="{class_name}" id="{item["html_id"]}">'
                f'<img src="{item["src"]}" alt="{item["caption"]}" />'
                f'<figcaption>图{item["number"]}&nbsp;&nbsp;{item["caption"]}{tag_html}</figcaption>'
                f'</figure>'
            )

        def _render_section_figures(section: str) -> str:
            return "".join(
                _render_standard_figure(item)
                for item in figure_manifest
                if item["section"] == section and not _is_pest_figure(item)
            )

        def _render_pest_gallery() -> str:
            cards = []
            for item in figure_manifest:
                if item["section"] != "insect" or not _is_pest_figure(item):
                    continue
                tag_html = (
                    f'&nbsp;<span class="gemini-tag">{item["tag"]}</span>'
                    if item.get("tag")
                    else ""
                )
                cards.append(
                    f'<figure class="pest-card" id="{item["html_id"]}">'
                    f'<img src="{item["src"]}" alt="{item["caption"]}" />'
                    f'<figcaption>图{item["number"]}&nbsp;&nbsp;{item["caption"]}{tag_html}</figcaption>'
                    f'</figure>'
                )
            if not cards:
                return ""
            return f'<div class="pest-gallery">{"".join(cards)}</div>'

        pest_gallery = _render_pest_gallery()

        methodology_html = ""
        if methodology:
            methodology_html = f"""
<div class="methodology-box">
  <div class="methodology-title">监测体系科学性说明</div>
  <p>{methodology.get('monitoring_statement', '')}</p>
  <p>{methodology.get('baseline_statement', '')}</p>
</div>"""

        history_compare_html = ""
        history_items = [
            item
            for key, item in (history_comparison.get("modules") or {}).items()
            if key != "spore"
        ]
        previous_period = history_comparison.get("previous_period", {}) or {}
        history_cards = []
        for item in history_items:
            unit = item.get("unit", "")
            history_cards.append(
                f"""
  <div class="guideline-card">
    <div class="guideline-label">{item.get('label', '—')}</div>
    <div class="guideline-value">{item.get('trend', '—')}</div>
    <div class="guideline-note">{item.get('metric_label', '—')}：本期 {_v(item.get('current_value'), f' {unit}' if unit else '')}，上一周期 {_v(item.get('previous_value'), f' {unit}' if unit else '')}，变化率 {_pct(item.get('change_rate'))}</div>
  </div>"""
            )
        if history_cards:
            history_compare_html = f"""
<div class="methodology-box">
  <div class="methodology-title">历史周期对比摘要</div>
  <p>对比口径：{history_comparison.get('comparison_basis', '本期与上一等长周期对比')}。上一周期区间为 {previous_period.get('start', '—')} 至 {previous_period.get('end', '—')}。</p>
</div>
<div class="guideline-grid">
{''.join(history_cards)}
</div>"""

        guideline_cards = [
            ("估算减蚀率", _pct(runoff_guideline.get("estimated_reduction_rate")), "监测型"),
            ("污染削减综合率", _pct(water_guideline.get("composite_reduction_rate")), "近30天"),
            ("病虫风险等级", pest_guideline.get("risk_level", "—"), "闭环研判"),
            ("气象支撑状态", water_source_support.get("status", "—"), "水源涵养"),
        ]
        guideline_overview_html = """
<div class="guideline-grid">
""" + "".join(
            f"""
  <div class="guideline-card">
    <div class="guideline-label">{label}</div>
    <div class="guideline-value">{value}</div>
    <div class="guideline-note">{note}</div>
  </div>"""
            for label, value, note in guideline_cards
        ) + """
</div>"""

        fixed_rules_html = ""
        confirmed_rules = implementation_matrix.get("confirmed_rules") or []
        if confirmed_rules:
            fixed_rules_html = """
<div class="methodology-box">
  <div class="methodology-title">本期已确认业务口径</div>
""" + "".join(
                f"  <p>{item}</p>" for item in confirmed_rules
            ) + """
</div>"""

        special_sections = build_special_analysis_sections(summary_dict)
        special_cards_html = "".join(
            f"""
  <article class="special-card">
    <div class="special-head">
      <h3>{item['title']}</h3>
      <span>{item.get('badge') or '专项研判'}</span>
    </div>
    <div class="special-facts">
      {''.join(f'<b>{fact}</b>' for fact in item.get('facts', []))}
    </div>
    {''.join(f'<p>{paragraph}</p>' for paragraph in item.get('paragraphs', []))}
    <div class="special-actions">
      <div class="special-actions-title">管理建议</div>
      {''.join(f'<em>{action}</em>' for action in item.get('actions', []))}
    </div>
  </article>"""
            for item in special_sections
        )

        weather_support_html = ""
        if weather.get("enabled") and weather.get("status") == "ok":
            current = weather.get("current", {}) or {}
            history_summary = weather.get("history_summary", {}) or {}
            history_range = weather.get("history_range", {}) or {}
            history_range_text = f"{history_range.get('start', '—')} 至 {history_range.get('end', '—')}"
            support_note_html = ""
            if water_source_support.get("message"):
                support_note_html = f'<div class="support-note">{water_source_support.get("message", "")}</div>'
            weather_support_html = f"""
<div class="support-grid">
  <div class="support-card">
    <div class="support-title">气象补充数据</div>
    <div class="support-row"><span>当前天气</span><strong>{current.get('text', '—')}</strong></div>
    <div class="support-row"><span>当前温度</span><strong>{_v(current.get('temp'), ' ℃')}</strong></div>
    <div class="support-row"><span>当前湿度</span><strong>{_v(current.get('humidity'), ' %')}</strong></div>
    <div class="support-row"><span>当前风速</span><strong>{_v(current.get('wind_speed'), ' km/h')}</strong></div>
  </div>
  <div class="support-card">
    <div class="support-title">水源涵养支撑说明</div>
    <div class="support-row"><span>历史区间</span><strong>{history_range_text}</strong></div>
    <div class="support-row"><span>最近7天累计降水</span><strong>{_v(history_summary.get('total_precip'), ' mm')}</strong></div>
    <div class="support-row"><span>最近7天平均气温</span><strong>{_v(history_summary.get('avg_temp_mean'), ' ℃')}</strong></div>
    <div class="support-row"><span>最近7天平均湿度</span><strong>{_v(history_summary.get('avg_humidity'), ' %')}</strong></div>
    <div class="support-row"><span>最近7天平均风速</span><strong>{_v(history_summary.get('avg_wind_speed'), ' km/h')}</strong></div>
    <div class="support-row"><span>降水日数</span><strong>{_v(history_summary.get('rainy_days'), ' 天')}</strong></div>
    {support_note_html}
  </div>
</div>"""

        runoff_rows = "".join(
            f"<tr><td>{item.get('name', item.get('device_code'))}</td>"
            f"<td class='num'>{_v(item.get('erosion_proxy'))}</td>"
            f"<td class='num'>{_v(item.get('avg_sand_content'))}</td>"
            f"<td class='num'>{_v(item.get('avg_runoff'))}</td>"
            f"<td class='num'>{_pct(item.get('relative_to_reference'))}</td></tr>"
            for item in runoff_guideline.get("station_metrics", [])
        )
        runoff_guideline_html = f"""
<div class="table-wrap">
  <table class="data-table">
    <caption>水土流失监测型估算结果（次生林参照）</caption>
    <thead><tr><th>监测点</th><th>侵蚀代理指标</th><th>平均含沙量</th><th>平均径流量</th><th>相对次生林差异</th></tr></thead>
    <tbody>{runoff_rows}</tbody>
  </table>
</div>""" if runoff_rows else ""

        water_rows = "".join(
            f"<tr><td>{item.get('label')}</td>"
            f"<td class='num'>{_v(item.get('baseline_avg'))}</td>"
            f"<td class='num'>{_v(item.get('recent_avg'))}</td>"
            f"<td class='num'>{_v(item.get('latest_value'))}</td>"
            f"<td class='num'>{_pct(item.get('recent_reduction_rate'))}</td>"
            f"<td class='num'>{_pct(item.get('latest_reduction_rate'))}</td></tr>"
            for item in water_guideline.get("metrics", [])
        )
        water_guideline_html = f"""
<div class="table-wrap">
  <table class="data-table">
    <caption>农业面源污染削减率（基准期对比）</caption>
    <thead><tr><th>指标</th><th>基准期平均</th><th>近30天平均</th><th>最新值</th><th>近30天削减率</th><th>最新削减率</th></tr></thead>
    <tbody>{water_rows}</tbody>
  </table>
</div>""" if water_rows else ""

        indicator_warnings = warning_analysis.get("indicator_warnings", []) or []
        warning_comparison = warning_analysis.get("comparison", {}) or {}

        def _render_warning_cards(items: list[dict[str, Any]], title: str, note: str = "") -> str:
            if not items:
                return ""
            cards = "".join(
                f"""
  <article class="warning-card warning-{item.get('level_code', 'normal')}">
    <div class="warning-top">
      <div>
        <div class="warning-label">{item.get('title', '—')}</div>
        <div class="warning-metric">{item.get('metric_label', '—')}</div>
      </div>
      <span class="warning-badge">{item.get('level', '—')}</span>
    </div>
    <div class="warning-value">{item.get('display_value', '—')}</div>
    <div class="warning-band">判定区间：{item.get('band', '—')}</div>
    <div class="warning-progress"><span style="width:{item.get('score', 0)}%"></span></div>
    <p class="warning-summary">{item.get('summary', '')}</p>
    <p class="warning-action">建议动作：{item.get('action', '—')}</p>
  </article>"""
                for item in items
            )
            note_html = f'<div class="warning-note">{note}</div>' if note else ""
            return f"""
<div class="warning-block">
  <div class="warning-block-title">{title}</div>
  {note_html}
  <div class="warning-grid">
{cards}
  </div>
</div>"""

        hydrology_warnings_html = _render_warning_cards(
            [
                item
                for item in indicator_warnings
                if item.get("key") in {"rainfall_peak", "sand_content"}
            ],
            "雨量与含沙分级预警",
            warning_comparison.get("message", ""),
        )
        pest_warnings_html = _render_warning_cards(
            [
                item
                for item in indicator_warnings
                if item.get("key") == "insect_peak"
            ],
            "虫情分级预警",
        )

        def _render_spore_image_appendix() -> str:
            images = (sp.get("capture_images") or [])[-12:]
            if not images:
                body = """
  <div class="ai-placeholder"><strong>孢子采集图像</strong>本监测周期未获取到孢子捕捉仪采集图像，当前保留附录位置占位。</div>"""
            else:
                cards = "".join(
                    f"""
  <figure class="chart-figure" id="fig-spore-appendix-{idx}">
    <img src="{img.get('url', '')}" alt="孢子采集实景 {idx}" />
    <figcaption>孢子采集实景 {idx}&nbsp;&nbsp;设备编号：{img.get('device_code', '—')}&nbsp;&nbsp;采集时间：{img.get('time', '—')}</figcaption>
  </figure>"""
                    for idx, img in enumerate(images, 1)
                    if img.get("url")
                )
                body = cards or """
  <div class="ai-placeholder"><strong>孢子采集图像</strong>本监测周期未获取到有效孢子图片地址。</div>"""

            return f"""
<section class="report-section" id="sec-spore-images">
  <h2 class="sec-title"><span class="sec-num">七</span>孢子采集图像附录</h2>
  {body}
</section>"""

        # ----------------------------------------------------------------
        # Build HTML sections
        # ----------------------------------------------------------------

        sec_hydrology = f"""
<section class="report-section" id="sec-hydrology">
  <h2 class="sec-title"><span class="sec-num">二</span>雨量与径流监测</h2>
  <div class="kpi-row">
    <div class="kpi-card kpi-blue">
      <div class="kpi-label">雨量记录</div>
      <div class="kpi-value">{_v(rn.get('records_count'), '', '0')}<span class="kpi-unit">条</span></div>
    </div>
    <div class="kpi-card kpi-blue">
      <div class="kpi-label">累计降雨量</div>
      <div class="kpi-value">{_v(rn.get('total_rainfall'), '', '0')}<span class="kpi-unit">mm</span></div>
    </div>
    <div class="kpi-card kpi-green">
      <div class="kpi-label">径流记录</div>
      <div class="kpi-value">{_v(ro.get('records_count'), '', '0')}<span class="kpi-unit">条</span></div>
    </div>
    <div class="kpi-card kpi-green">
      <div class="kpi-label">监测点数量</div>
      <div class="kpi-value">{_v(ro.get('device_count'), '', '0')}<span class="kpi-unit">个</span></div>
    </div>
    <div class="kpi-card">
      <div class="kpi-label">累计径流量</div>
      <div class="kpi-value">{_v(ro.get('total_runoff'), '', '0')}<span class="kpi-unit">m3</span></div>
    </div>
  </div>
  {weather_support_html}
  {hydrology_warnings_html}
  {_render_section_figures('hydrology')}
  {runoff_guideline_html}
</section>"""

        sec_water_quality = f"""
<section class="report-section" id="sec-water-quality">
  <h2 class="sec-title"><span class="sec-num">三</span>水质监测</h2>
  <div class="kpi-row">
    <div class="kpi-card">
      <div class="kpi-label">水质记录</div>
      <div class="kpi-value">{_v(wq.get('records_count'), '', '0')}<span class="kpi-unit">条</span></div>
    </div>
    <div class="kpi-card kpi-blue">
      <div class="kpi-label">平均氨氮</div>
      <div class="kpi-value">{_v(wq.get('avg_nh3_n'))}<span class="kpi-unit">mg/L</span></div>
    </div>
    <div class="kpi-card kpi-green">
      <div class="kpi-label">平均总磷</div>
      <div class="kpi-value">{_v(wq.get('avg_tp'))}<span class="kpi-unit">mg/L</span></div>
    </div>
    <div class="kpi-card kpi-warm">
      <div class="kpi-label">平均高猛酸盐</div>
      <div class="kpi-value">{_v(wq.get('avg_permanganate'))}<span class="kpi-unit">mg/L</span></div>
    </div>
    <div class="kpi-card kpi-warn">
      <div class="kpi-label">平均总氮</div>
      <div class="kpi-value">{_v(wq.get('avg_tn'))}<span class="kpi-unit">mg/L</span></div>
    </div>
  </div>
  {_render_section_figures('water_quality')}
  {water_guideline_html}
</section>"""

        species_rows = "".join(
            f"<tr><td>{item[0]}</td><td class='num'>{item[1]}</td>"
            f"<td class='num'>{item[1] / max(ins.get('total_count', 1), 1) * 100:.1f}%</td></tr>"
            for item in (ins.get("top_species") or [])
        )
        species_table = f"""
<div class="table-wrap">
  <table class="data-table">
    <caption>主要虫种捕获统计</caption>
    <thead><tr><th>虫种名称</th><th>捕获数量（只）</th><th>占比</th></tr></thead>
    <tbody>{species_rows}</tbody>
  </table>
</div>""" if species_rows else ""

        sec_insect = f"""
<section class="report-section" id="sec-insect">
  <h2 class="sec-title"><span class="sec-num">四</span>虫情测报监测</h2>
  <div class="kpi-row">
    <div class="kpi-card">
      <div class="kpi-label">监测记录</div>
      <div class="kpi-value">{_v(ins.get('records_count'), '', '0')}<span class="kpi-unit">条</span></div>
    </div>
    <div class="kpi-card kpi-warn">
      <div class="kpi-label">期间捕获总数</div>
      <div class="kpi-value">{_v(ins.get('total_count', 0))}<span class="kpi-unit">只</span></div>
    </div>
    <div class="kpi-card">
      <div class="kpi-label">记录虫种数</div>
      <div class="kpi-value">{len(ins.get('top_species') or [])}<span class="kpi-unit">种</span></div>
    </div>
  </div>
  {_render_section_figures('insect')}
  {pest_warnings_html}
  {species_table}
  {pest_gallery}
</section>"""

        sec_special = f"""
<section class="report-section special-section" id="sec-special">
  <h2 class="sec-title"><span class="sec-num">五</span>四类深度专项分析</h2>
  <div class="special-intro">
    围绕虫情、雨情、水土流失与径流、面源水质污染四个专项方向，结合本期监测统计、历史同口径对比和管理阈值进行分项研判，形成可直接用于巡查、预警和处置的专项结论。
  </div>
  <div class="special-grid">
    {special_cards_html}
  </div>
</section>"""

        # Section 7: AI
        if ai_html:
            ai_body = f'<div class="ai-body">{ai_html}</div>'
        else:
            ai_body = '<div class="ai-placeholder"><strong>AI 智能分析</strong>（配置 DEEPSEEK_API_KEY 后自动生成完整分析报告）</div>'

        sec_ai = f"""
<section class="report-section ai-section" id="sec-ai">
  <h2 class="sec-title"><span class="sec-num">六</span>AI 综合分析报告
    <span class="ai-badge">DeepSeek</span>
  </h2>
  {ai_body}
</section>"""

        sec_spore_images = _render_spore_image_appendix()

        html = f"""<!DOCTYPE html>
<html lang="zh-CN">
<head>
  <meta charset="UTF-8" />
  <meta name="viewport" content="width=device-width, initial-scale=1.0" />
  <title>三亚市天涯区橡胶林近自然化改造和农田提升监测平台 — {report_type}</title>
  <style>
/* ============================================================
   全局重置与基础
   ============================================================ */
*, *::before, *::after {{ box-sizing: border-box; margin: 0; padding: 0; }}

body {{
  font-family: "Microsoft YaHei", "微软雅黑", "PingFang SC", "Helvetica Neue", sans-serif;
  font-size: 14px;
  line-height: 1.8;
  background: #F4F6F9;
  color: #2C3E50;
  min-height: 100vh;
}}

/* ============================================================
   封面页
   ============================================================ */
.cover {{
  background: linear-gradient(160deg, #0d2137 0%, #0a3d62 55%, #104e8b 100%);
  color: #fff;
  min-height: 320px;
  padding: 60px 80px 50px;
  position: relative;
  overflow: hidden;
  page-break-after: always;
}}
.cover::before {{
  content: "";
  position: absolute;
  inset: 0;
  background:
    radial-gradient(circle at 15% 40%, rgba(255,255,255,0.04) 0%, transparent 50%),
    radial-gradient(circle at 85% 20%, rgba(88,166,255,0.08) 0%, transparent 40%);
  pointer-events: none;
}}
.cover-platform {{
  font-size: 13px;
  letter-spacing: 4px;
  color: #79c0ff;
  text-transform: uppercase;
  margin-bottom: 20px;
  position: relative;
}}
.cover-title {{
  font-size: 36px;
  font-weight: 900;
  letter-spacing: 2px;
  line-height: 1.25;
  color: #fff;
  position: relative;
  margin-bottom: 12px;
}}
.cover-subtitle {{
  font-size: 18px;
  font-weight: 400;
  color: #a5d6ff;
  letter-spacing: 2px;
  position: relative;
  margin-bottom: 36px;
}}
.cover-meta {{
  display: flex;
  gap: 32px;
  flex-wrap: wrap;
  position: relative;
}}
.cover-meta-item {{
  background: rgba(255,255,255,0.08);
  border: 1px solid rgba(255,255,255,0.15);
  border-radius: 8px;
  padding: 10px 20px;
}}
.cover-meta-label {{
  font-size: 11px;
  color: #79c0ff;
  letter-spacing: 2px;
  margin-bottom: 4px;
}}
.cover-meta-value {{
  font-size: 15px;
  font-weight: 700;
  color: #e6edf3;
  letter-spacing: 1px;
}}
.cover-divider {{
  position: absolute;
  bottom: 0;
  left: 0;
  right: 0;
  height: 4px;
  background: linear-gradient(90deg, #1f6aa5, #58a6ff, #1f6aa5);
}}

/* ============================================================
   目录
   ============================================================ */
.toc-section {{
  background: #fff;
  max-width: 900px;
  margin: 32px auto 0;
  border-radius: 12px;
  box-shadow: 0 2px 16px rgba(0,0,0,0.07);
  padding: 32px 40px;
}}
.toc-title {{
  font-size: 16px;
  font-weight: 700;
  color: #1a5276;
  border-bottom: 2px solid #2E86C1;
  padding-bottom: 10px;
  margin-bottom: 18px;
  letter-spacing: 2px;
}}
.toc-list {{
  list-style: none;
}}
.toc-list li {{
  display: flex;
  justify-content: space-between;
  align-items: baseline;
  padding: 6px 0;
  border-bottom: 1px dashed #DDDDDD;
  font-size: 14px;
}}
.toc-list li:last-child {{ border-bottom: none; }}
.toc-list a {{
  color: #2C3E50;
  text-decoration: none;
  font-weight: 500;
}}
.toc-list a:hover {{ color: #2E86C1; }}
.toc-num {{
  color: #2E86C1;
  font-weight: 700;
  margin-right: 8px;
  min-width: 24px;
  display: inline-block;
}}

/* ============================================================
   报告主体
   ============================================================ */
.report-body {{
  max-width: 900px;
  margin: 28px auto 60px;
  padding: 0 16px;
}}

/* ============================================================
   章节
   ============================================================ */
.report-section {{
  background: #fff;
  border-radius: 12px;
  box-shadow: 0 2px 16px rgba(0,0,0,0.07);
  margin-bottom: 28px;
  overflow: hidden;
}}

.sec-title {{
  font-size: 18px;
  font-weight: 800;
  color: #fff;
  background: linear-gradient(90deg, #1a5276, #2E86C1);
  padding: 16px 28px;
  letter-spacing: 1px;
  display: flex;
  align-items: center;
  gap: 10px;
}}
.sec-num {{
  display: inline-flex;
  align-items: center;
  justify-content: center;
  width: 28px;
  height: 28px;
  border-radius: 50%;
  background: rgba(255,255,255,0.25);
  font-size: 13px;
  font-weight: 900;
  flex-shrink: 0;
}}

/* ============================================================
   概述卡片（Section 1）
   ============================================================ */
.overview-section .sec-title {{
  background: linear-gradient(90deg, #1a5276, #2471a3);
}}
.overview-grid {{
  display: grid;
  grid-template-columns: repeat(auto-fit, minmax(180px, 1fr));
  gap: 1px;
  background: #eee;
  margin: 0;
}}
.overview-item {{
  background: #fff;
  padding: 20px 24px;
  text-align: center;
}}
.overview-label {{
  font-size: 11px;
  color: #7F8C8D;
  letter-spacing: 1px;
  margin-bottom: 8px;
  text-transform: uppercase;
}}
.overview-value {{
  font-size: 28px;
  font-weight: 900;
  color: #2C3E50;
}}
.overview-unit {{
  font-size: 13px;
  color: #7F8C8D;
  margin-left: 2px;
  font-weight: 400;
}}

.methodology-box {{
  margin: 18px 24px 10px;
  padding: 16px 18px;
  border-radius: 10px;
  border: 1px solid #DCEAF5;
  background: linear-gradient(180deg, #F8FCFF 0%, #F2F7FB 100%);
}}
.methodology-title {{
  font-size: 14px;
  font-weight: 800;
  color: #1A5276;
  margin-bottom: 8px;
}}
.methodology-box p {{
  font-size: 13px;
  color: #36506A;
  line-height: 1.9;
  margin-bottom: 6px;
}}
.methodology-box p:last-child {{
  margin-bottom: 0;
}}

.guideline-grid {{
  display: grid;
  grid-template-columns: repeat(auto-fit, minmax(160px, 1fr));
  gap: 12px;
  padding: 10px 24px 20px;
}}
.guideline-card {{
  border-radius: 10px;
  border: 1px solid #E2ECF3;
  background: #fff;
  padding: 14px 16px;
}}
.guideline-label {{
  font-size: 11px;
  color: #6C7A89;
  margin-bottom: 6px;
  letter-spacing: 1px;
}}
.guideline-value {{
  font-size: 24px;
  font-weight: 800;
  color: #1A5276;
  line-height: 1.2;
}}
.guideline-note {{
  margin-top: 6px;
  font-size: 11px;
  color: #7F8C8D;
}}

.support-grid {{
  display: grid;
  grid-template-columns: repeat(auto-fit, minmax(260px, 1fr));
  gap: 14px;
  padding: 18px 24px 6px;
}}
.support-card {{
  border: 1px solid #E2ECF3;
  border-radius: 10px;
  background: #FBFDFF;
  padding: 16px;
}}
.support-title {{
  font-size: 14px;
  font-weight: 800;
  color: #1A5276;
  margin-bottom: 10px;
}}
.support-row {{
  display: flex;
  justify-content: space-between;
  gap: 12px;
  padding: 6px 0;
  border-bottom: 1px dashed #E4EBF1;
  font-size: 13px;
  color: #35516B;
}}
.support-row:last-of-type {{
  border-bottom: none;
}}
.support-row strong {{
  color: #1F4E79;
}}
.support-note {{
  margin-top: 10px;
  font-size: 12px;
  line-height: 1.8;
  color: #536B82;
}}

.warning-block {{
  padding: 0 24px 20px;
}}
.warning-block-title {{
  font-size: 14px;
  font-weight: 800;
  color: #1A5276;
  margin-bottom: 10px;
}}
.warning-note {{
  margin-bottom: 12px;
  padding: 10px 12px;
  border-radius: 10px;
  background: #F8FBFE;
  border: 1px dashed #D7E7F2;
  color: #4B6278;
  font-size: 12px;
  line-height: 1.8;
}}
.warning-grid {{
  display: grid;
  grid-template-columns: repeat(auto-fit, minmax(260px, 1fr));
  gap: 14px;
}}
.warning-card {{
  border-radius: 12px;
  border: 1px solid #E2ECF3;
  background: linear-gradient(180deg, #FFFFFF 0%, #F8FBFE 100%);
  padding: 16px;
  box-shadow: 0 8px 20px rgba(26, 82, 118, 0.06);
}}
.warning-top {{
  display: flex;
  justify-content: space-between;
  gap: 12px;
  align-items: flex-start;
}}
.warning-label {{
  font-size: 14px;
  font-weight: 800;
  color: #17324D;
}}
.warning-metric {{
  margin-top: 4px;
  font-size: 12px;
  color: #6B7C8D;
  line-height: 1.6;
}}
.warning-badge {{
  flex-shrink: 0;
  border-radius: 999px;
  padding: 5px 10px;
  font-size: 12px;
  font-weight: 800;
  line-height: 1;
}}
.warning-value {{
  margin-top: 14px;
  font-size: 28px;
  font-weight: 900;
  color: #17324D;
  line-height: 1.1;
}}
.warning-band {{
  margin-top: 8px;
  font-size: 12px;
  color: #5D7185;
}}
.warning-progress {{
  height: 8px;
  margin-top: 12px;
  border-radius: 999px;
  background: #EAF1F6;
  overflow: hidden;
}}
.warning-progress span {{
  display: block;
  height: 100%;
  border-radius: inherit;
  background: linear-gradient(90deg, #7DD3FC 0%, #1D4ED8 100%);
}}
.warning-summary,
.warning-action {{
  margin-top: 10px;
  font-size: 13px;
  line-height: 1.85;
  color: #35516B;
}}
.warning-action {{
  color: #1F4E79;
  font-weight: 600;
}}
.warning-attention {{
  border-color: #FDE68A;
}}
.warning-attention .warning-badge {{
  background: #FEF3C7;
  color: #92400E;
}}
.warning-severe {{
  border-color: #FDA4AF;
}}
.warning-severe .warning-badge {{
  background: #FFE4E6;
  color: #BE123C;
}}
.warning-high {{
  border-color: #FDBA74;
}}
.warning-high .warning-badge {{
  background: #FFEDD5;
  color: #C2410C;
}}
.warning-critical {{
  border-color: #FCA5A5;
}}
.warning-critical .warning-badge {{
  background: #FEE2E2;
  color: #B91C1C;
}}

/* ============================================================
   KPI 卡片行
   ============================================================ */
.kpi-row {{
  display: flex;
  flex-wrap: wrap;
  gap: 12px;
  padding: 20px 24px;
  background: #F8FAFC;
  border-bottom: 1px solid #EFEFEF;
}}
.kpi-card {{
  background: #fff;
  border: 1px solid #E8EDF2;
  border-radius: 10px;
  padding: 14px 20px;
  min-width: 120px;
  flex: 1;
  border-top: 3px solid #2E86C1;
}}
.kpi-card.kpi-warm  {{ border-top-color: #E67E22; }}
.kpi-card.kpi-green {{ border-top-color: #28B463; }}
.kpi-card.kpi-blue  {{ border-top-color: #2980B9; }}
.kpi-card.kpi-warn  {{ border-top-color: #C0392B; }}
.kpi-card.kpi-purple {{ border-top-color: #8E44AD; }}

.kpi-label {{
  font-size: 11px;
  color: #7F8C8D;
  letter-spacing: 1px;
  margin-bottom: 6px;
}}
.kpi-value {{
  font-size: 22px;
  font-weight: 800;
  color: #2C3E50;
  line-height: 1;
}}
.kpi-unit {{
  font-size: 12px;
  color: #7F8C8D;
  font-weight: 400;
  margin-left: 2px;
}}

/* ============================================================
   图表
   ============================================================ */
.chart-figure {{
  margin: 0 auto;
  padding: 24px 24px 16px;
  border-bottom: 1px solid #F0F0F0;
  display: flex;
  flex-direction: column;
  align-items: center;
}}
.chart-figure img {{
  width: 100%;
  max-width: 860px;
  height: auto;
  display: block;
  border-radius: 8px;
  border: 1px solid #E8EDF2;
}}
.chart-figure figcaption {{
  margin-top: 10px;
  font-size: 13px;
  color: #555;
  text-align: center;
  letter-spacing: 0.5px;
  font-weight: 500;
  width: 100%;
}}
.ai-img-figure img {{
  max-width: 720px;
  border: 1px solid #D5E8F3;
  box-shadow: 0 2px 12px rgba(0,0,0,0.08);
}}

/* ============================================================
   数据表格
   ============================================================ */
.table-wrap {{
  padding: 0 24px 20px;
  overflow-x: auto;
}}
.data-table {{
  width: 100%;
  border-collapse: collapse;
  font-size: 14px;
  margin-top: 12px;
}}
.data-table caption {{
  font-weight: 700;
  color: #1a5276;
  text-align: left;
  padding: 8px 0 6px;
  font-size: 14px;
  letter-spacing: 0.5px;
}}
.data-table th {{
  background: #1a5276;
  color: #fff;
  font-weight: 700;
  padding: 10px 14px;
  text-align: left;
  font-size: 13px;
  letter-spacing: 0.5px;
}}
.data-table td {{
  padding: 10px 14px;
  border-bottom: 1px solid #EFEFEF;
  color: #2C3E50;
  vertical-align: middle;
  line-height: 1.7;
}}
.data-table td.num {{
  text-align: right;
  font-weight: 600;
  font-variant-numeric: tabular-nums;
}}
.data-table tr:nth-child(even) td {{ background: #F8FAFC; }}
.data-table tr:hover td {{ background: #EBF5FB; }}

/* ============================================================
   四类深度专项分析
   ============================================================ */
.special-section .sec-title {{
  background: linear-gradient(90deg, #14532D, #15803D);
}}
.special-intro {{
  margin: 20px 24px 0;
  padding: 14px 16px;
  border-radius: 10px;
  border: 1px solid #D8EFE0;
  background: #F6FCF8;
  color: #315545;
  font-size: 13px;
  line-height: 1.9;
}}
.special-grid {{
  display: grid;
  grid-template-columns: 1fr;
  gap: 16px;
  padding: 20px 24px 24px;
}}
.special-card {{
  border: 1px solid #DDEBE3;
  border-radius: 12px;
  background: linear-gradient(180deg, #FFFFFF 0%, #FAFDFB 100%);
  padding: 18px;
  box-shadow: 0 8px 22px rgba(20, 83, 45, 0.06);
}}
.special-head {{
  display: flex;
  justify-content: space-between;
  align-items: flex-start;
  gap: 12px;
  margin-bottom: 12px;
}}
.special-head h3 {{
  font-size: 16px;
  color: #174D32;
  line-height: 1.35;
}}
.special-head span {{
  flex-shrink: 0;
  padding: 5px 10px;
  border-radius: 999px;
  background: #E8F7EE;
  color: #166534;
  border: 1px solid #BFE7CC;
  font-size: 12px;
  font-weight: 800;
}}
.special-facts {{
  display: grid;
  grid-template-columns: repeat(auto-fit, minmax(150px, 1fr));
  gap: 8px;
  margin-bottom: 12px;
}}
.special-facts b {{
  display: block;
  border-radius: 9px;
  border: 1px solid #E2ECF3;
  background: #F8FBFE;
  color: #1F4E79;
  font-size: 12px;
  line-height: 1.5;
  padding: 8px 10px;
}}
.special-card p {{
  margin: 8px 0;
  color: #344E45;
  font-size: 13.5px;
  line-height: 1.95;
  text-align: justify;
}}
.special-actions {{
  margin-top: 12px;
  display: grid;
  gap: 8px;
}}
.special-actions-title {{
  font-size: 13px;
  font-weight: 800;
  color: #14532D;
}}
.special-actions em {{
  display: block;
  border-left: 3px solid #22C55E;
  background: #F4FBF6;
  color: #355B48;
  font-size: 13px;
  line-height: 1.8;
  font-style: normal;
  padding: 8px 10px;
}}

/* ============================================================
   AI 分析区
   ============================================================ */
.ai-section .sec-title {{
  background: linear-gradient(90deg, #1a3a4a, #1f6aa5);
}}
.ai-badge {{
  margin-left: auto;
  background: rgba(255,255,255,0.15);
  border: 1px solid rgba(255,255,255,0.3);
  border-radius: 20px;
  padding: 3px 12px;
  font-size: 11px;
  font-weight: 400;
  letter-spacing: 1px;
}}
.ai-body {{
  padding: 28px 36px;
  color: #2C3E50;
  font-size: 14.5px;
  line-height: 2;
}}
.ai-body h2.ai-h2 {{
  font-size: 17px;
  font-weight: 800;
  color: #1a5276;
  margin: 28px 0 12px;
  padding-bottom: 8px;
  border-bottom: 2px solid #D6EAF8;
  letter-spacing: 1px;
}}
.ai-body h2.ai-h2:first-child {{ margin-top: 0; }}
.ai-body h3.ai-h3 {{
  font-size: 15px;
  font-weight: 700;
  color: #2E86C1;
  margin: 20px 0 8px;
  padding-left: 10px;
  border-left: 4px solid #2E86C1;
  letter-spacing: 0.5px;
}}
.ai-body p {{
  margin-bottom: 12px;
  text-align: justify;
  text-indent: 2em;
}}
.ai-body li {{
  margin: 6px 0 6px 24px;
  list-style-type: disc;
}}
.ai-body strong {{
  font-weight: 700;
  color: #1a5276;
}}
.ai-placeholder {{
  padding: 40px;
  text-align: center;
  color: #7F8C8D;
  font-size: 14px;
}}
.ai-placeholder strong {{
  display: block;
  font-size: 16px;
  color: #2C3E50;
  margin-bottom: 8px;
}}

/* ============================================================
   页脚
   ============================================================ */
.page-footer {{
  background: #1a5276;
  color: rgba(255,255,255,0.6);
  text-align: center;
  padding: 20px 40px;
  font-size: 12px;
  letter-spacing: 1px;
  margin-top: 40px;
}}
.page-footer strong {{
  color: rgba(255,255,255,0.9);
  font-weight: 600;
}}

/* ============================================================
   Gemini AI 图片样式
   ============================================================ */
.ai-img-figure img {{
  max-width: 720px;
  width: 100%;
  max-height: none;
  aspect-ratio: 16 / 9;
  object-fit: cover;
  border: 1px solid #D5E8F3;
  box-shadow: 0 2px 12px rgba(0,0,0,0.08);
}}
.gemini-tag {{
  display: inline-block;
  background: linear-gradient(90deg, #4285F4, #EA4335, #FBBC04, #34A853);
  -webkit-background-clip: text;
  -webkit-text-fill-color: transparent;
  font-weight: 700;
  font-size: 11px;
}}

/* 虫种图鉴 Gallery */
.pest-gallery {{
  display: flex;
  flex-wrap: wrap;
  gap: 16px;
  padding: 20px 24px;
  background: #F8FAFC;
  border-top: 1px solid #EFEFEF;
}}
.pest-card {{
  flex: 1;
  min-width: 220px;
  max-width: 320px;
  background: #fff;
  border: 1px solid #E8EDF2;
  border-radius: 10px;
  overflow: hidden;
  text-align: center;
  box-shadow: 0 2px 8px rgba(0,0,0,0.06);
  display: flex;
  flex-direction: column;
  align-items: center;
}}
.pest-card img {{
  width: 100%;
  aspect-ratio: 16 / 9;
  object-fit: cover;
  display: block;
}}
.pest-card figcaption {{
  font-size: 12px;
  color: #555;
  text-align: center;
  padding: 8px 8px 10px;
  font-weight: 500;
  line-height: 1.4;
}}
.pest-card .pest-name {{
  font-size: 13px;
  font-weight: 700;
  color: #2C3E50;
  padding: 8px 8px 2px;
}}
.pest-card .pest-label {{
  font-size: 10px;
  color: #7F8C8D;
  padding-bottom: 8px;
  letter-spacing: 0.5px;
}}

/* ============================================================
   打印优化
   ============================================================ */
@media print {{
  body {{ background: #fff; }}
  .report-section, .toc-section {{ box-shadow: none; border: 1px solid #ddd; }}
  .sec-title {{ -webkit-print-color-adjust: exact; print-color-adjust: exact; }}
  .kpi-card {{ border-top-width: 3px; -webkit-print-color-adjust: exact; print-color-adjust: exact; }}
}}
  </style>
</head>
<body>

<!-- ============================================================
     封面
     ============================================================ -->
<div class="cover" {cover_style}>
  <div class="cover-platform">三亚市天涯区橡胶林近自然化改造和农田提升监测平台</div>
  <div class="cover-title">橡胶林近自然化改造生态效益评估报告</div>
  <div class="cover-subtitle">{report_type}</div>
  <div class="cover-meta">
    <div class="cover-meta-item">
      <div class="cover-meta-label">监测起始</div>
      <div class="cover-meta-value">{start_str}</div>
    </div>
    <div class="cover-meta-item">
      <div class="cover-meta-label">监测截止</div>
      <div class="cover-meta-value">{end_str}</div>
    </div>
    <div class="cover-meta-item">
      <div class="cover-meta-label">报告生成时间</div>
      <div class="cover-meta-value">{generated_at}</div>
    </div>
    <div class="cover-meta-item">
      <div class="cover-meta-label">智能分析</div>
      <div class="cover-meta-value">DeepSeek · Gemini</div>
    </div>
  </div>
  <div class="cover-divider"></div>
</div>

<!-- ============================================================
     目录
     ============================================================ -->
<div class="toc-section">
  <div class="toc-title">目 &nbsp; 录</div>
  <ul class="toc-list">
    <li><a href="#sec-overview"><span class="toc-num">一、</span>监测期综合概况与评估背景</a></li>
    <li><a href="#sec-hydrology"><span class="toc-num">二、</span>雨量与径流监测</a></li>
    <li><a href="#sec-water-quality"><span class="toc-num">三、</span>水质监测</a></li>
    <li><a href="#sec-insect"><span class="toc-num">四、</span>虫情测报监测</a></li>
    <li><a href="#sec-special"><span class="toc-num">五、</span>四类深度专项分析</a></li>
    <li><a href="#sec-ai"><span class="toc-num">六、</span>AI 综合分析报告</a></li>
    <li><a href="#sec-spore-images"><span class="toc-num">七、</span>孢子采集图像附录</a></li>
  </ul>
</div>

<!-- ============================================================
     正文
     ============================================================ -->
<div class="report-body">

  <!-- 一、数据概览 -->
  <section class="report-section overview-section" id="sec-overview">
    <h2 class="sec-title"><span class="sec-num">一</span>监测期综合概况与评估背景</h2>
    <div class="overview-grid">
      <div class="overview-item">
        <div class="overview-label">水质记录</div>
        <div class="overview-value">{_v(summary_dict.get('water_quality', {}).get('records_count'), '', '0')}<span class="overview-unit">条</span></div>
      </div>
      <div class="overview-item">
        <div class="overview-label">虫情捕获</div>
        <div class="overview-value">{_v(ins.get('total_count', 0))}<span class="overview-unit">只</span></div>
      </div>
      <div class="overview-item">
        <div class="overview-label">记录虫种</div>
        <div class="overview-value">{len(ins.get('top_species') or [])}<span class="overview-unit">种</span></div>
      </div>
      <div class="overview-item">
        <div class="overview-label">累计降雨</div>
        <div class="overview-value">{_v(summary_dict.get('rain', {}).get('total_rainfall'), '', '0')}<span class="overview-unit">mm</span></div>
      </div>
    </div>
    {methodology_html}
    {history_compare_html}
  </section>

  {sec_hydrology}
  {sec_water_quality}
  {sec_insect}
  {sec_special}
  {sec_ai}
  {sec_spore_images}

</div><!-- /report-body -->

<footer class="page-footer">
  <strong>三亚市天涯区橡胶林近自然化改造和农田提升监测平台</strong>&nbsp;·&nbsp;
  数据来源：橡胶林生态监测设备网络&nbsp;·&nbsp;
  报告生成时间：{generated_at}&nbsp;·&nbsp;
  AI 分析由 DeepSeek 提供支持
</footer>

</body>
</html>"""
        return html
